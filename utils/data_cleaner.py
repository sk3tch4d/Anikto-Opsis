# ==============================
# DATA_CLEANER.PY — CLEAN XLSX FILES FOR ROUTING
# ==============================

import pandas as pd
import numpy as np
import os
import logging
import tempfile
import threading
import time
import json
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

from .data_format import format_fillrate, format_cart_ops

# ==============================
# IMPORT RENAMES & REMOVE
# ==============================
def load_json(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logging.error(f"[CONFIG] Failed to load {path}: {e}")
        return {}  # or [] if it's for REMOVE_COLUMNS

CONFIG_DIR = os.path.join(os.path.dirname(__file__), "configs")
COLUMN_RENAMES = load_json(os.path.join(CONFIG_DIR, "column_renames.json"))
REMOVE_COLUMNS = load_json(os.path.join(CONFIG_DIR, "remove_columns.json"))

logging.debug(f"[CONFIG] Loaded COLUMN_RENAMES: {len(COLUMN_RENAMES)} entries")
logging.debug(f"[CONFIG] Loaded REMOVE_COLUMNS: {len(REMOVE_COLUMNS)} columns")

# ==============================
# LOG CLEANING
# ==============================
def log_cleaning(step, df, extra=""):
    name = df.attrs.get("name", "Unnamed")
    logging.debug(f"[CLEAN]🧹 {step} -> {name}{f' — {extra}' if extra else ''}")

# ==============================
# DETECT AND SET HEADER
# ==============================
def detect_and_set_header(df, max_rows=20):
    for i in range(max_rows):
        row = df.iloc[i].fillna("").astype(str).str.strip()
        num_strings = sum(val != "" for val in row)
        unique_values = len(set(val for val in row if val != ""))

        # Debug: log each candidate row
        logging.debug(f"[CLEAN]🧪 Row {i}: {num_strings} non-empty strings, {unique_values} unique values")
        logging.debug(f"[CLEAN]🧪 Row {i} contents: {row.tolist()}")

        if num_strings >= len(row) // 3 and unique_values >= 2:
            df.columns = row
            df = df.iloc[i + 1:].reset_index(drop=True)

            logging.debug(f"[CLEAN]🧹 Detected header at row {i}")
            logging.debug(f"[CLEAN]🧹 Set columns: {df.columns.tolist()}")
            return df

    logging.debug("[CLEAN]🧹 No valid header row detected in preview window.")
    return df

# ==============================
# DETECT UNION VALUE
# ==============================
def detect_union_value(df):
    if any(str(col).strip().upper() == "UNION" for col in df.iloc[0]):
        log_cleaning("Union Column Present — Skipping Detection", df)
        return None

    preview = df.head(5).fillna("").astype(str).apply(lambda x: x.str.upper())
    text = " ".join(preview.values.flatten())
    logging.debug(f"[DEBUG] Preview flattened text: {text!r}")

    if "OPSEU" in text:
        log_cleaning("Union Found:", df, extra="OPSEU")
        return "OPSEU"
    elif "CUPE" in text:
        log_cleaning("Union Found:", df, extra="CUPE")
        return "CUPE"

    log_cleaning("No Union Found:", df)
    return None

# ==============================
# CLEANING FUNCTIONS (STEP MODULES)
# ==============================
def clean_headers(df):
    df.columns = df.columns.str.strip().str.replace(r"\s+", " ", regex=True)
    rename_map = {k: v for k, v in COLUMN_RENAMES.items() if k in df.columns}
    df.rename(columns=rename_map, inplace=True)
    log_cleaning("Headers", df)
    return df
# ==============================
def clean_columns(df):
    df = df.loc[:, ~df.columns.duplicated()]
    df.drop(columns=[col for col in REMOVE_COLUMNS if col in df.columns], inplace=True, errors='ignore')
    log_cleaning("Columns", df, extra=f"{len(df.columns)} columns remain")
    return df
# ==============================
def clean_deleted_rows(df):
    mask = df.astype(str).apply(lambda x: x.str.contains('DELETED', case=False, na=False)).any(axis=1)
    log_cleaning("Deleted Rows", df, extra=f"{mask.sum()} rows removed")
    return df[~mask]
# ==============================
def clean_flags(df):
    for col in ['Mat', 'Pl', 'D', 'Del']:
        if col in df.columns:
            df = df[df[col].astype(str).str.upper() != 'X']
    log_cleaning("Flags", df)
    return df
# ==============================
def clean_format(df):
    sort_cols = []
    if 'Material' in df.columns: sort_cols.append('Material')
    if 'Bin' in df.columns: sort_cols.append('Bin')
    #if 'USL' in df.columns: sort_cols.append('USL')
    for col in sort_cols:
        df[col] = df[col].astype(str)  # Convert to string 
    if sort_cols:
        df = df.sort_values(by=sort_cols, ascending=True, na_position='last')
    log_cleaning("Formatting", df)
    return df
# ==============================
def clean_lint(df):
    filler_pattern = r"[-=~*#_.]{3,}"
    lint_mask = df.apply(lambda row: row.astype(str).str.fullmatch(filler_pattern)).any(axis=1)
    log_cleaning("Lint Rows", df, extra=f"{lint_mask.sum()} rows removed")
    return df[~lint_mask]

# ==============================
# XLSX CLEANING PIPELINE
# ==============================
def clean_xlsx(file_stream, *steps, header=None, name=None, detect_header=True, multi_sheet=True, format=False):
    if multi_sheet:
        sheet_dict = pd.read_excel(file_stream, sheet_name=None, header=None)
        cleaned_dfs = []

        for sheet_name, df in sheet_dict.items():
            df.attrs["name"] = sheet_name

            # Strip column names and drop empty rows
            df.columns = [str(col).strip() for col in df.columns]
            df = df.dropna(how="all")

            # Detect union BEFORE header stripping
            union = detect_union_value(df)

            if detect_header:
                df = detect_and_set_header(df)

            log_cleaning("Cleaning Sheet", df, extra=f"Sheet: {sheet_name}")

            df = clean_lint(df)
            
            # Apply all cleaning steps
            for step in steps:
                df = step(df)

            # Normalize dates
            for col in ['Created', 'Date', 'First']:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors='coerce').dt.date
                    log_cleaning("Normalized Date", df)

            # Attach union value AFTER cleaning
            if union:
                df["Union"] = [union] * len(df)
                log_cleaning("Detected Union", df, extra=union)

            cleaned_dfs.append(df)

        if not cleaned_dfs:
            return pd.DataFrame()

        df_combined = pd.concat(cleaned_dfs, ignore_index=True)

        # if name and "clean" in name.lower():
        if format == True:
            df_combined = format_cart_ops(df_combined)
            df_combined = format_fillrate(df_combined)

        df_combined.attrs["name"] = name or "Combined Sheets"
        return df_combined

    else:
        # Load only the first sheet
        xls = pd.ExcelFile(file_stream)
        sheet_name = xls.sheet_names[0]
        df = pd.read_excel(file_stream, header=None, sheet_name=sheet_name)
        df.attrs["name"] = name or sheet_name

        # Detect union BEFORE header stripping
        union = detect_union_value(df)

        if detect_header:
            df = detect_and_set_header(df)

        df.columns = [str(col).strip() for col in df.columns]
        df = df.dropna(how="all")

        log_cleaning("Cleaning Sheet", df, extra=f"Sheet: {sheet_name}")

        df = clean_lint(df)
        
        for step in steps:
            df = step(df)

        for col in ['Created', 'Date', 'First']:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.date
                log_cleaning("Normalized Date", df)

        # Attach union value AFTER cleaning
        if union:
            df["Union"] = [union] * len(df)
            log_cleaning("Detected Union", df, extra=union)

        # if name and "clean" in name.lower():
        if format == True:
            df = format_cart_ops(df)
            df = format_fillrate(df)

        return df

# ==============================
# DB CLEANING PIPELINE
# ==============================
def clean_db(df, name="DB Inventory"):
    #df = detect_and_set_header(df) # Adjust Header if needed
    df.attrs["name"] = name
    steps = [clean_headers, clean_columns, clean_deleted_rows, clean_flags, clean_format]
    for step in steps:
        df = step(df)
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.date
        log_cleaning("Normalized Date", df)
    return df

# ==============================
# EXCEL COLUMN AUTO-FIT
# ==============================
def autofit_columns(worksheet, max_width=40, min_width=4, padding=2):
    for col_cells in worksheet.columns:
        lengths = [len(str(cell.value)) if cell.value else 0 for cell in col_cells]
        best_fit = min(max(max(lengths) + padding, min_width), max_width)
        col_letter = get_column_letter(col_cells[0].column)
        worksheet.column_dimensions[col_letter].width = best_fit

# ==============================
# SAVE TO TEMP FILE FOR DOWNLOAD
# ==============================
def save_cleaned_df(df, filename=None):

    if filename:
        path = os.path.join("/tmp", filename)
    else:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir="/tmp") as tmp:
            path = tmp.name

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
        worksheet = writer.sheets["Sheet1"]

        # Auto-fit columns
        autofit_columns(worksheet)

        # Freeze header row
        worksheet.freeze_panes = worksheet["A2"]

        # Apply borders to all cells with data
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                       min_col=1, max_col=worksheet.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = border

    log_cleaning("Saved File", df)
    return path

# ==============================
# SCHEDULE FILE DELETION
# ==============================
def schedule_file_deletion(path, delay_seconds=600):
    def delete_file():
        try:
            if os.path.exists(path):
                os.remove(path)
                logging.info(f"[CLEANUP] Deleted expired file: {path}")
        except Exception as e:
            logging.warning(f"[CLEANUP] Failed to delete {path}: {e}")

    threading.Timer(delay_seconds, delete_file).start()
