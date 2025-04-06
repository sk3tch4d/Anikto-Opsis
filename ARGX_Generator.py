
import os
import re
import pdfplumber
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule

# === Constants ===
VALID_NAMES = {
    "Adeniyi, Oluwaseyi", "Bhardwaj, Liam", "Donovan, Patrick", "Gallivan, David",
    "Janaway, Alexander", "Robichaud, Richard", "Santo, Jaime", "Tobin, James",
    "Ukwesa, Jennifer", "Vanderputten, Richard", "Woodland, Nathaniel"
}
PAY_PERIOD_START_DATE = datetime(2025, 1, 13)

# === Core Helpers ===
def get_pay_period(date_obj):
    return (date_obj - PAY_PERIOD_START_DATE.date()).days // 14

def extract_shift_ids(line):
    return [t for t in line.split() if re.match(r'(SA[1-4]|[A-Z]*\d{3,4})$', t, re.IGNORECASE)]

def classify_shift(start, end, shift_ids):
    if "313" in shift_ids:
        return "Day"
    s = datetime.strptime(start, "%H:%M").time()
    e = datetime.strptime(end, "%H:%M").time()
    if s >= datetime.strptime("07:00", "%H:%M").time() and e <= datetime.strptime("15:00", "%H:%M").time():
        return "Day"
    if s >= datetime.strptime("15:00", "%H:%M").time() and e <= datetime.strptime("23:00", "%H:%M").time():
        return "Evening"
    if s >= datetime.strptime("23:00", "%H:%M").time() or e <= datetime.strptime("07:00", "%H:%M").time():
        return "Night"
    if (s >= datetime.strptime("08:00", "%H:%M").time() and e <= datetime.strptime("12:00", "%H:%M").time()):
        return "Day"
    return "Other"

# === PDF Parser ===
def parse_pdf(pdf_path):
    records = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            lines = text.splitlines() if text else []
            current_date = None
            for line in lines:
                if "Inventory Services" in line:
                    try:
                        current_date = datetime.strptime(line.split()[-1], "%d/%b/%Y").date()
                    except:
                        continue
                    continue
                if any(x in line for x in ["Off:", "On Call", "Relief"]):
                    continue
                tokens = line.strip().split()
                if len(tokens) >= 5:
                    try:
                        start_time = tokens[-4]
                        end_time = tokens[-3]
                        full_name = f"{tokens[-2].rstrip(',')}, {tokens[-1]}"
                        if full_name not in VALID_NAMES:
                            continue
                        shift_ids = extract_shift_ids(line)
                        full_shift_id = " ".join(shift_ids).strip()
                        dt_start = datetime.strptime(f"{current_date} {start_time}", "%Y-%m-%d %H:%M")
                        dt_end = datetime.strptime(f"{current_date} {end_time}", "%Y-%m-%d %H:%M")
                        if dt_end <= dt_start:
                            dt_end += timedelta(days=1)
                        hours = round((dt_end - dt_start).seconds / 3600, 1)
                        shift_type = classify_shift(start_time, end_time, shift_ids)
                        records.append({
                            "Name": full_name,
                            "Date": current_date.strftime("%a, %b %d"),
                            "DateObj": current_date,
                            "Shift": full_shift_id,
                            "Type": shift_type,
                            "Hours": hours,
                            "Start": start_time,
                            "End": end_time
                        })
                    except:
                        continue
    return pd.DataFrame(records)


# === Excel Writer ===
def write_argx_v2(df, output_path):
    wb = Workbook()
    bold = Font(bold=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    medium_bottom = Border(bottom=Side(style="medium"))
    all_weeks = sorted(df["DateObj"].apply(lambda d: d - timedelta(days=d.weekday())).unique())
    df["WeekStart"] = df["DateObj"].apply(lambda d: d - timedelta(days=d.weekday()))
    names = sorted(df["Name"].unique())

    # === Weekly Totals Sheet ===
    ws_totals = wb.active
    ws_totals.title = "Weekly Totals"
    ws_totals.freeze_panes = "A2"
    ws_totals.column_dimensions['A'].width = 24.0
    for i in range(len(all_weeks)):
        col_letter = chr(66 + i)
        ws_totals.column_dimensions[col_letter].width = 13.0

    ws_totals.cell(row=1, column=1, value="Name").font = bold
    ws_totals.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    for i, week in enumerate(all_weeks, start=2):
        cell = ws_totals.cell(row=1, column=i, value=week.strftime("%Y-%m-%d"))
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    pivot = df.groupby(["Name", "WeekStart"])["Hours"].sum().unstack(fill_value=0)
    for row_idx, name in enumerate(names, start=2):
        ws_totals.cell(row=row_idx, column=1, value=name).alignment = Alignment(horizontal="center")
        ws_totals.cell(row=row_idx, column=1).border = thin
        for col_idx, week in enumerate(all_weeks, start=2):
            hours = round(pivot.loc[name, week], 1) if week in pivot.columns else 0.0
            cell = ws_totals.cell(row=row_idx, column=col_idx, value=hours)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin

    for i, week in enumerate(all_weeks, start=2):
        if get_pay_period(week) % 2 == 1:
            col_letter = chr(64 + i)
            ws_totals.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(names)+1}",
                FormulaRule(formula=["TRUE"], fill=PatternFill(fill_type="solid", fgColor="FFD9D9D9")))

    # === Employee Sheets ===
    headers = ["Date", "Shift", "Type", "Hours", "Start", "End"]
    aligns = ["left", "center", "left", "center", "center", "center"]
    widths = [13.0, 7.0, 9.0, 8.0, 8.0, 8.0]

    for name, group in df.groupby("Name"):
        sheetname = " ".join(name.replace(",", "").split()[::-1])
        ws = wb.create_sheet(title=sheetname)
        ws.freeze_panes = "A2"
        for col, (head, align, width) in enumerate(zip(headers, aligns, widths), start=1):
            cell = ws.cell(row=1, column=col, value=head)
            cell.font = bold
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
            ws.column_dimensions[chr(64 + col)].width = width

        group = group.sort_values("DateObj").reset_index(drop=True)
        for row_idx, row in group.iterrows():
            cur_period = get_pay_period(row["DateObj"])
            values = [row["Date"], row["Shift"], row["Type"], row["Hours"], row["Start"], row["End"]]
            for col, (val, align) in enumerate(zip(values, aligns), start=1):
                cell = ws.cell(row=row_idx + 2, column=col, value=val)
                cell.alignment = Alignment(horizontal=align)
                if cur_period % 2 == 1:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
            if row_idx + 1 < len(group):
                next_period = get_pay_period(group.loc[row_idx + 1, "DateObj"])
                if next_period != cur_period:
                    for col in range(1, 7):
                        ws.cell(row=row_idx + 2, column=col).border = medium_bottom

    wb.save(output_path)

# === Generate ARGX ===
def generate_argx_from_pdfs(pdf_paths, output_xlsx, log_duplicates=True):
    frames = [parse_pdf(p) for p in pdf_paths]
    df = pd.concat(frames, ignore_index=True)

    print(f"Total parsed rows: {len(df)}")
    print(f"PDF paths: {pdf_paths}")

    
    if df.empty:
        print("No data found.")
        return None

    if log_duplicates:
        dups = df[df.duplicated(subset=["Name", "DateObj", "Shift"], keep="first")]
        if not dups.empty:
            dups.to_excel("ARGX_DroppedDuplicates_Log.xlsx", index=False)

    df = df.drop_duplicates(subset=["Name", "DateObj", "Shift"])
    write_argx_v2(df, output_xlsx)
    print(f"Saved: {output_xlsx}")
    return output_xlsx

# === Generate ARGM ===
def generate_heatmap_png(df, date_label):
    df["WeekStart"] = df["DateObj"].apply(lambda d: d - timedelta(days=d.weekday()))

    pivot = df.pivot_table(index="Name", columns="WeekStart", values="Hours", aggfunc="sum", fill_value=0)
    pivot = pivot.round(0).astype(int)
    plt.figure(figsize=(10, 6))
    sns.heatmap(pivot, annot=True, fmt="d", cmap="Blues")
    path = f"/tmp/ARGM_{date_label}.png"
    plt.title("Weekly Hours per Person")
    plt.tight_layout()
    plt.savefig(path)
    plt.close()
    print(f"Saved heatmap: {path}")
    return path
    
# === Compatibility alias ===
def generate_argx_and_heatmap(pdf_paths, generate_argx=True, generate_heatmap=False):
    frames = [parse_pdf(p) for p in pdf_paths]
    df = pd.concat(frames, ignore_index=True)

    if df.empty:
        print("No data found.")
        return None

    df = df.drop_duplicates(subset=["Name", "DateObj", "Shift"])
    first_date = df["DateObj"].min().strftime("%Y-%m-%d")
    output_files = []

    if generate_argx:
        output_filename = f"ARGX_{first_date}.xlsx"
        output_path = os.path.join("/tmp", output_filename)
        write_argx_v2(df, output_path)
        print(f"Saved: {output_path}")
        output_files.append(output_path)

    if generate_heatmap:
        heatmap_path = generate_heatmap_png(df, first_date)  # Assuming this exists
        output_files.append(heatmap_path)

    from collections import defaultdict
def group_by_shift(df, target_date):
        shifts = defaultdict(list)
        for _, row in df[df["DateObj"] == target_date].sort_values("Name").iterrows():
            shifts[row["Type"]].append((row["Name"], row["Shift"]))
        return dict(shifts)

        today = datetime.now().date()
        tomorrow = today + timedelta(days=1)

        stats = {
            "working_today": group_by_shift(df, today),
            "working_tomorrow": group_by_shift(df, tomorrow),
            "total_hours_week": round(df[df["WeekStart"] == today - timedelta(days=today.weekday())]["Hours"].sum()),
            "top_day": df.groupby("DateObj")["Hours"].sum().idxmax(),
            "top_day_hours": int(df.groupby("DateObj")["Hours"].sum().max()),
            "rankings": df.groupby("Name")["Hours"].sum().sort_values(ascending=False).astype(int).items()
        }

        return output_files, stats
    return output_files