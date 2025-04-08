from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
from datetime import timedelta, datetime
import pandas as pd


def make_pay_period_fn(start_date):
    """
    Creates a pay period calculation function based on a specific start date.

    Parameters:
        start_date (datetime): The start date of the first pay period.

    Returns:
        function: A function that takes a date object and returns the number
                  of 14-day pay periods since the start_date.
    """
    def get_pay_period(date_obj):
        return (date_obj - start_date.date()).days // 14
    return get_pay_period


def write_argx(df, output_path, get_pay_period):
    """
    Writes shift data to an Excel workbook with weekly totals and individual sheets per person.

    Parameters:
        df (pd.DataFrame): DataFrame containing parsed shift data.
        output_path (str): File path for the resulting Excel file.
        get_pay_period (function): A function that calculates pay period number from a date.
    """
    wb = Workbook()
    bold = Font(bold=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    medium_bottom = Border(bottom=Side(style="medium"))
    all_weeks = sorted(df["DateObj"].apply(lambda d: d - timedelta(days=d.weekday())).unique())
    df["WeekStart"] = df["DateObj"].apply(lambda d: d - timedelta(days=d.weekday()))
    names = sorted(df["Name"].unique())

    # === Weekly Totals Sheet ===
    ws_totals = wb.active
    ws_totals.title = "Weekly Totals"
    ws_totals.freeze_panes = "A2"
    ws_totals.column_dimensions['A'].width = 24.0
    for i in range(len(all_weeks)):
        col_letter = chr(66 + i)
        ws_totals.column_dimensions[col_letter].width = 13.0

    ws_totals.cell(row=1, column=1, value="Name").font = bold
    ws_totals.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    for i, week in enumerate(all_weeks, start=2):
        cell = ws_totals.cell(row=1, column=i, value=week.strftime("%Y-%m-%d"))
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    pivot = df.groupby(["Name", "WeekStart"])["Hours"].sum().unstack(fill_value=0)
    for row_idx, name in enumerate(names, start=2):
        ws_totals.cell(row=row_idx, column=1, value=name).alignment = Alignment(horizontal="center")
        ws_totals.cell(row=row_idx, column=1).border = thin
        for col_idx, week in enumerate(all_weeks, start=2):
            hours = round(pivot.loc[name, week], 1) if week in pivot.columns else 0.0
            cell = ws_totals.cell(row=row_idx, column=col_idx, value=hours)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin

    for i, week in enumerate(all_weeks, start=2):
        if get_pay_period(week) % 2 == 1:
            col_letter = chr(64 + i)
            ws_totals.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(names)+1}",
                FormulaRule(formula=["TRUE"], fill=PatternFill(fill_type="solid", fgColor="FFD9D9D9")))

    # === Employee Sheets ===
    headers = ["Date", "Shift", "Type", "Hours", "Start", "End"]
    aligns = ["left", "center", "left", "center", "center", "center"]
    widths = [13.0, 7.0, 9.0, 8.0, 8.0, 8.0]

    for name, group in df.groupby("Name"):
        sheetname = " ".join(name.replace(",", "").split()[::-1])
        ws = wb.create_sheet(title=sheetname)
        ws.freeze_panes = "A2"
        for col, (head, align, width) in enumerate(zip(headers, aligns, widths), start=1):
            cell = ws.cell(row=1, column=col, value=head)
            cell.font = bold
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
            ws.column_dimensions[chr(64 + col)].width = width

        group = group.sort_values("DateObj").reset_index(drop=True)
        for row_idx, row in group.iterrows():
            cur_period = get_pay_period(row["DateObj"])
            values = [row["Date"], row["Shift"], row["Type"], row["Hours"], row["Start"], row["End"]]
            for col, (val, align) in enumerate(zip(values, aligns), start=1):
                cell = ws.cell(row=row_idx + 2, column=col, value=val)
                cell.alignment = Alignment(horizontal=align)
                if cur_period % 2 == 1:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
            if row_idx + 1 < len(group):
                next_period = get_pay_period(group.loc[row_idx + 1, "DateObj"])
                if next_period != cur_period:
                    for col in range(1, 7):
                        ws.cell(row=row_idx + 2, column=col).border = medium_bottom

    wb.save(output_path)
